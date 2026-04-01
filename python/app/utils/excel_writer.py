"""
Excel writer utility for exporting orders
"""
import asyncio
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.order import Order


class ExcelWriter:
    """Excel writer for order exports"""

    # Excel headers mapping
    HEADERS = [
        "订单ID",
        "订单编号",
        "用户ID",
        "用户姓名",
        "用户手机号",
        "用户身份证",
        "用户邮箱",
        "用户地址",
        "商品ID",
        "商品名称",
        "商品分类",
        "商品单价",
        "购买数量",
        "订单总金额",
        "优惠金额",
        "实付金额",
        "订单状态",
        "支付方式",
        "支付时间",
        "订单来源",
        "收货地址",
        "收货人姓名",
        "收货人电话",
        "物流单号",
        "发货时间",
        "完成时间",
        "备注",
        "创建时间",
        "更新时间",
    ]

    FIELDS = [
        "order_id",
        "order_no",
        "user_id",
        "user_name",
        "user_phone",
        "user_id_card",
        "user_email",
        "user_address",
        "product_id",
        "product_name",
        "product_category",
        "product_price",
        "quantity",
        "total_amount",
        "discount_amount",
        "pay_amount",
        "order_status",
        "payment_method",
        "payment_time",
        "order_source",
        "shipping_address",
        "receiver_name",
        "receiver_phone",
        "logistics_no",
        "delivery_time",
        "complete_time",
        "remark",
        "created_at",
        "updated_at",
    ]

    # Column widths (approximate)
    COLUMN_WIDTHS = {
        "订单ID": 12,
        "订单编号": 20,
        "用户ID": 12,
        "用户姓名": 15,
        "用户手机号": 15,
        "用户身份证": 20,
        "用户邮箱": 25,
        "用户地址": 30,
        "商品ID": 12,
        "商品名称": 25,
        "商品分类": 15,
        "商品单价": 12,
        "购买数量": 12,
        "订单总金额": 15,
        "优惠金额": 15,
        "实付金额": 15,
        "订单状态": 12,
        "支付方式": 12,
        "支付时间": 20,
        "订单来源": 15,
        "收货地址": 30,
        "收货人姓名": 15,
        "收货人电话": 15,
        "物流单号": 20,
        "发货时间": 20,
        "完成时间": 20,
        "备注": 30,
        "创建时间": 20,
        "更新时间": 20,
    }

    @staticmethod
    def format_value(value) -> Optional[str]:
        """Format a value for Excel output"""
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(value, Decimal):
            return float(value)
        return str(value)

    @classmethod
    def _setup_workbook(cls, workbook: Workbook) -> None:
        """Setup workbook styles and formatting"""
        # Get active worksheet
        ws = workbook.active
        ws.title = "订单数据"

        # Define styles
        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        data_font = Font(name="微软雅黑", size=10)
        data_alignment = Alignment(horizontal="left", vertical="center")

        # Write headers
        for col_idx, header in enumerate(cls.HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Set column widths
        for col_idx, header in enumerate(cls.HEADERS, start=1):
            width = cls.COLUMN_WIDTHS.get(header, 15)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze first row
        ws.freeze_panes = "A2"

    @classmethod
    async def write_to_file(
        cls,
        orders: list[Order],
        file_path: Path,
        batch_size: int = 1000,
    ) -> int:
        """
        Write orders to Excel file asynchronously

        Args:
            orders: List of order objects
            file_path: Output file path
            batch_size: Number of records to process in each batch

        Returns:
            Number of records written
        """
        # Create parent directory if needed
        file_path.parent.mkdir(parents=True, exist_ok=True)

        # Create workbook
        workbook = Workbook()
        ws = workbook.active

        # Setup workbook
        cls._setup_workbook(workbook)

        # Define data styles
        data_font = Font(name="微软雅黑", size=10)
        data_alignment = Alignment(horizontal="left", vertical="center")

        # Write data in batches
        row_idx = 2
        count = 0

        for i in range(0, len(orders), batch_size):
            batch = orders[i : i + batch_size]

            for order in batch:
                for col_idx, field in enumerate(cls.FIELDS, start=1):
                    value = getattr(order, field)
                    formatted_value = cls.format_value(value)

                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = formatted_value
                    cell.font = data_font
                    cell.alignment = data_alignment

                row_idx += 1
                count += 1

            # Yield control to event loop
            await asyncio.sleep(0)

        # Save workbook
        workbook.save(file_path)

        return count

    @classmethod
    async def generate_excel_bytes(cls, orders: list[Order]) -> bytes:
        """
        Generate Excel content as bytes

        Args:
            orders: List of order objects

        Returns:
            Excel content as bytes
        """
        from io import BytesIO

        # Create workbook
        workbook = Workbook()

        # Setup workbook
        cls._setup_workbook(workbook)
        ws = workbook.active

        # Define data styles
        data_font = Font(name="微软雅黑", size=10)
        data_alignment = Alignment(horizontal="left", vertical="center")

        # Write data
        row_idx = 2
        for order in orders:
            for col_idx, field in enumerate(cls.FIELDS, start=1):
                value = getattr(order, field)
                formatted_value = cls.format_value(value)

                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = formatted_value
                cell.font = data_font
                cell.alignment = data_alignment

            row_idx += 1

            # Yield control to event loop periodically
            if row_idx % 100 == 0:
                await asyncio.sleep(0)

        # Save to bytes
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()
